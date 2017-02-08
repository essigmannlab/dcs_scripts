#!/usr/bin/env python3

import os
import argparse
import matplotlib as mpl
import matplotlib.pyplot as plt

from Bio import SeqIO
from openpyxl import Workbook
from collections import OrderedDict
from itertools import cycle, product
from matplotlib.patches import Rectangle
from Bio.Seq import Seq, reverse_complement
from openpyxl.styles import Font, colors, PatternFill

purines, pyrimidines = ('A', 'G'), ('C', 'T')
dna_bases = sorted(purines + pyrimidines)

py_labels = ('C>A', 'C>G', 'C>T', 'T>A', 'T>C', 'T>G')
pu_labels = ('A>C', 'A>G', 'A>T', 'G>A', 'G>C', 'G>T')

sig_colors = ['#52C3F1', '#231F20', '#E62223', '#CBC9C8', '#97D54C', '#EDBFC2']


def axes_off(ax):
    ax.get_xaxis().set_visible(False)
    ax.get_yaxis().set_visible(False)
    return(ax)


def cleanup_chart_junk(ax):
    ax.grid(False)
    ax.patch.set_facecolor('white')
    ax.yaxis.set_ticks_position('left')
    ax.xaxis.set_ticks_position('bottom')
    ax.get_xaxis().tick_bottom()
    ax.get_yaxis().tick_left()
    for tic in ax.yaxis.get_major_ticks() + ax.xaxis.get_major_ticks():
        tic.tick2On = False
    return ax


def despine(ax):
    for spine in ['top', 'left', 'bottom', 'right']:
        ax.spines[spine].set_visible(False)
    return ax


def ticks_off(ax):
    for tic in ax.yaxis.get_major_ticks() + ax.xaxis.get_major_ticks():
        tic.tick1On = tic.tick2On = False
    return ax


class Mutation(object):
    def __init__(self, ref, base, chrom, position, context):
        self.ref = ref
        self.base = base
        self.chrom = chrom
        self.position = position
        self.context = context

    @property
    def substitution(self):
        return Seq(''.join([self.ref, '>', self.base]))

    @property
    def context(self):
        return Seq(''.join([self._5prime, self.ref, self._3prime]))

    @context.setter
    def context(self, context):
        self._5prime = context[:len(context) - 2]
        self._3prime = context[len(context) - 1:]

    @property
    def label(self):
        return ''.join([self._5prime.upper(),
                        '[',
                        str(self.substitution).upper(),
                        ']',
                        self._3prime.upper()])

    def __repr__(self):
        return 'Mutation: {}\nPosition: {}:{}'.format(str(self.substitution),
                                                      self.chrom,
                                                      self.position)


class Spectrum(OrderedDict):
    def __init__(self, notation, kmer=3):
        super(Spectrum, self).__init__()
        self.notation = notation
        self.kmer = kmer

        if self.kmer < 1 or (self.kmer % 2) != 1:
            raise ValueError('kmer must be a positive odd integer')
        if notation not in ['pyrimidine', 'purine']:
            raise ValueError('Notation must be "pyrimidine" or "purine"')

        substitutions = py_labels if notation == 'pyrimidine' else pu_labels

        for substitution in substitutions:
            for context in dna_kmers(self.kmer):
                if context[int((self.kmer - 1) / 2)] == substitution[0]:
                    self[substitution, context] = 0

    @property
    def variant_total(self):
        return sum(self.values())

    @property
    def proportion(self):
        if self.variant_total > 0:
            return [n / self.variant_total for n in self.values()]
        else:
            return list(self.values())


def colored_bins(division, ax=None, colors=None, labels=None, padding=0):
    if ax is None:
        ax = plt.gca()

    ax = axes_off(despine(ticks_off(cleanup_chart_junk(ax))))

    colors = cycle(['0.8']) if colors is None else cycle(colors)

    for bin in range(division):
        ax.add_patch(Rectangle(xy=(bin / division, 0),
                               width=(1 / division) - padding,
                               height=1,
                               color=next(colors)))

    if labels is not None:
        ax.get_xaxis().set_visible(True)
        xticks = [(x - 0.5) / division for x in range(1, division + 1)]
        ax.set_xticks(xticks)
        ax.set_xticklabels(labels)
    return ax


def dna_kmers(k=3):
    """
    Returns the cartesian product (unique) list of all DNA kmers of
    length k.

    Parameters
    ----------
    k : int
        Length of DNA kmers.

    Returns
    -------
    generator : generator of strings
        Alphabetically sorted cartesian product of all DNA kmers of
        length k.
    """
    for prod in product(dna_bases, repeat=k):
        yield ''.join(prod)


def fasta_to_dict(ref_file):

    with open(ref_file, 'r') as handle:
        return SeqIO.to_dict(SeqIO.parse(handle, 'fasta'))


def from_mutpos(mutpos_file, ref_file, clonality=(0, 1), min_depth=100, kmer=3,
                chromosome=None, start=0, end=None, notation='pyrimidine',
                fmt='essigmann', verbose=False):
    mutations = []
    record_dict = fasta_to_dict(ref_file)

    # Open the mutpos_file and read data line by line
    with open(mutpos_file, 'r') as handle:
        for line in handle:
            # Strip newline characters and split on tabs
            line = line.strip().split('\t')

            # Unpack line and cast to proper data type
            chrom, ref = str(line[0]), str(line[1]).upper()
            position, depth = int(line[2]) - 1, int(line[3])

            if chromosome is not None and chrom != chromosome:
                continue
            if position < start:
                continue
            if end is not None and position >= end:
                continue

            if fmt == 'essigmann':
                A, C, G, T, N = map(int, line[4:9])
            elif fmt == 'wesdirect':
                # N = 0
                thisMut = {base: 0 for base in dna_bases}
                thisMut[str(line[5])] = int(line[4])
                A = thisMut['A']
                C = thisMut['C']
                G = thisMut['G']
                T = thisMut['T']
            elif fmt == 'loeb':
                T, C, G, A = map(int, line[5:9])
                # N = int(line[11])
            else:
                raise ValueError('Format must be essigmann, loeb, wesdirect')

            # If we see no observed base substitutions, skip this loop
            if sum([A, C, G, T]) == 0:
                continue

            # Read mapped depth (minus Ns) at this position must be greater
            # than min_depth, if not, skip this loop
            if min_depth > depth:
                continue

            # Get the kmer context at the given position in the given
            # chromosome as looked up in record_dict. If we encounter an
            # edge case (return of None) we'll skip this loop
            context = str(get_kmer(record_dict, chrom, position, kmer)).upper()

            if context is None:
                continue

            # If the base is not in our intended labeling set: let's get
            # the complement of the reference base, the reverse complement
            # of the trinucleotide context, and the mutation counts for the
            # complement of the base substitution observed
            if (
                (notation == 'purine' and ref in pyrimidines) or
                (notation == 'pyrimidine' and ref in purines)
            ):
                ref = str(reverse_complement(ref))
                context = str(reverse_complement(context))
                A, G, C, T = T, C, G, A

            for base, num_mutations in zip(dna_bases, [A, C, G, T]):
                # Clonality is defined as the frequency of any base
                # substitution at this one genomic position. For example, if
                # there are 5% T, 10% C, and 100% G in a reference position of
                # A, we can seletion a clonality filter of (0.1, 0.5) to
                # eliminate the rare T mutations and clonal G mutations.
                base_clonality = num_mutations / depth
                if not min(clonality) <= base_clonality <= max(clonality):
                    continue

                for _ in range(num_mutations):
                    mutation = Mutation(ref, base, chrom, position, context)
                    mutation.depth, mutation.clonality = depth, base_clonality
                    mutations.append(mutation)

    if verbose is True:
        print('Found {} Mutations'.format(len(mutations)))
    return mutations


def get_kmer(record_dict, chromosome, position, k=3, pos='mid'):
    """
    Given a dictionary (in memory) of fasta sequences this function will
    return a kmer of length k centered about pos at a certain genomic position
    in an identified dictionary key i.e. chromosome.

    Parameters
    ----------
    record_dict : dict of Bio.Seq objects
        Dictionary where keys are chromosomes and values are Bio.Seq
        sequences.
    chromosome : str
        Key to use when accessing the record_dict.
    position : int
        Position in sequence to query.
    k : int
        Odd number for length of kmer.
    pos : 'mid' or int
        Position in kmer that the genomic position should be centered on.

    Returns
    -------
    kmer : str
        kmer of length k centered around position in chromosome
    """

    if pos == 'mid' and k % 2 != 1:
        raise ValueError("Even length DNA has no midpoint")

    if pos == 'mid':
        pos = int((k - 1) / 2)

    assert k > pos, "Cannot index past length of k"

    start = position - pos
    end = position + (k - pos)

    try:
        chromosome_length = len(str(record_dict[chromosome].seq))
    except ValueError:
        raise ValueError(
            'Chromosome {} not found in reference'.format(chromosome))

    # It is necessary to protect for the two edge cases now
    if start >= 0 and end <= chromosome_length:
        try:
            return str(record_dict[chromosome].seq[start:end])
        except ValueError:
            ValueError(
                'Position {} not found in chromosome {}'.format(
                    start, chromosome))
    else:
        return None


def spectrum(heights, ax=None, xlabels=None, y_max=None, **kwargs):

    if ax is None:
        ax = plt.gca()
    ax = ticks_off(cleanup_chart_junk(ax))
    ax.yaxis.grid(True, color=str(0.8), ls='-')

    ax.set_xlim([-0.5, len(heights)])
    if y_max is not None:
        ax.set_ylim(ax.get_ylim()[0], y_max)

    bar_width = kwargs.pop('bar_width', 0.65)

    bars = ax.bar(left=range(len(heights)),
                  height=heights,
                  width=bar_width,
                  zorder=3)

    for i, color in enumerate([c for c in sig_colors for _ in range(16)]):
        bars[i].set_color(color)

    ax.set_xticks([tick + bar_width / 2 for tick in range(len(heights))])
    ax.set_xticklabels(xlabels, family='monospace', rotation=90)
    ax.set_ylabel(kwargs.pop('ylabel', 'Percent of Mutations'))
    ax.set_title(kwargs.pop('title', None), y=0.84)
    return ax


def spectrum_map(nrow, ncol, heights, xlabels=None, labels=None, y_max=None,
                 titles=None, x_inches=16, ylabel=None):

    aspect = 4 / 11
    fig, axes = plt.subplots(nrow * 2, ncol,
                             figsize=(x_inches * ncol,
                                      x_inches * nrow * aspect),
                             gridspec_kw={'height_ratios': nrow * [28, 1],
                                          'hspace': 0.2,
                                          'wspace': 0.07})

    axes_iter = iter(enumerate(axes.flatten()))

    for column in range(ncol * 2):
        if column % 2 == 0:
            for i, ax in [next(axes_iter) for _ in range(nrow)]:
                if i >= len(heights) * 2:
                    ax = axes_off(ax)
                    continue

                spectrum(
                    heights[i],
                    ax=ax,
                    y_max=y_max,
                    ylabel=ylabel,
                    xlabels=None if xlabels is None else xlabels[i],
                    title=None if titles is None else titles[i])
        else:
            for i, ax in [next(axes_iter) for _ in range(nrow)]:
                if i >= len(heights) * 2:
                    ax = axes_off(ax)
                    continue

                colored_bins(
                    division=6,
                    ax=ax,
                    colors=sig_colors,
                    labels=labels)
    return fig, axes


def main():
    info = ("Find the non-normalized trinucleotide mutation frequencies from "
            "parsing a FASTA reference and .mutpos file as outputted from the"
            "duplex sequencing pipeline.")
    parser = argparse.ArgumentParser(description=info)
    parser.add_argument("-r", "--ref_file",
                        action="store",
                        type=str,
                        dest="ref_file",
                        help="The reference genome in FASTA format.",
                        required=True)
    parser.add_argument("-m", "--mutpos_file",
                        action="store",
                        type=str,
                        dest="mutpos_file",
                        help="The mutpos file.",
                        required=True)
    parser.add_argument("-c", "--minClonality",
                        action="store",
                        type=float,
                        dest="min_clonality",
                        default=0,
                        help=("The minimum clonality for counting a mutation "
                              "[0]"),
                        required=False)
    parser.add_argument("-C", "--maxClonality",
                        action="store",
                        type=float,
                        dest="max_clonality",
                        default=0.2,
                        help=("The maximum clonality for counting a mutation "
                              "(inclusive) [0.2]"),
                        required=False)
    parser.add_argument("-d", "--minDepth",
                        action="store",
                        type=int,
                        dest="min_depth",
                        default=100,
                        help=("The minimum depth of reads for each location "
                              "(inclusive) [100]."),
                        required=False)
    parser.add_argument("-n", "--notation",
                        action="store",
                        type=str,
                        dest="notation",
                        default="pyrimidine",
                        help=("This is useful for labelling figures as"
                              "purines or pyrimidines. [pyrimidine]"),
                        required=False)
    parser.add_argument("-i", "--dpi",
                        action="store",
                        type=int,
                        dest="dpi",
                        default=320,
                        help=("The dots per inch of the final saved figure "
                              "[320]."),
                        required=False)
    parser.add_argument("-f", "--format",
                        action="store",
                        type=str,
                        dest="format",
                        default='png',
                        help=("The format of the output image file. Popular "
                              "options are png or svg [png]."),
                        required=False)
    parser.add_argument("-y", "--ymax",
                        action="store",
                        dest="ymax",
                        default=None,
                        help="Sets the Y_MAX value in figure. Must be 0-100.",
                        required=False)
    parser.add_argument("-s", "--save",
                        type=str,
                        dest="save",
                        default='both',
                        help=("Whether to save both, ratio, or total "
                              "[both]."),
                        required=False)
    parser.add_argument("-t", "--title",
                        type=str,
                        dest="title",
                        default='',
                        help=("Plot title to be overlayed if supplied. The "
                              "default option creates an informative title. "
                              "Type None for no title."),
                        required=False)
    parser.add_argument("--lab",
                        type=str,
                        dest="lab",
                        default="loeb",
                        help=("Mutpos file format (loeb, essigmann, "
                              "wesdirect) [loeb]."),
                        required=False)

    args = parser.parse_args()

    # Set up local variables
    mpl.rc("savefig", dpi=int(args.dpi))
    clonality = (args.min_clonality, args.max_clonality)

    mutpos_name = os.path.basename(args.mutpos_file)

    if args.mutpos_file.endswith('.mutpos'):
        image_file1 = mutpos_name.replace('.mutpos',
                                          '-ratio.' + args.format)
        image_file2 = mutpos_name.replace('.mutpos',
                                          '-total.' + args.format)
        data_file = mutpos_name.replace('.mutpos', '.xlsx')
    else:
        image_file1 = mutpos_name + '-ratio.' + args.format
        image_file2 = mutpos_name + '-total.' + args.format
        data_file = mutpos_name + '.xlsx'

    # Parse mutpos file
    mutations = from_mutpos(args.mutpos_file,
                            args.ref_file,
                            clonality=clonality,
                            min_depth=args.min_depth,
                            notation=args.notation,
                            fmt=args.lab,
                            verbose=True)

    data = Spectrum(notation=args.notation, kmer=3)

    for mutation in mutations:
        data[str(mutation.substitution), str(mutation.context)] += 1

    # Format and save to Excel
    wb = Workbook()
    ws = wb.active
    ws.append(['Substitution', 'Context', 'Count', 'Proportion'])

    for letter in 'ABCD':
        ws[letter + '1'].fill = PatternFill(start_color=colors.BLACK,
                                            end_color=colors.BLACK,
                                            fill_type='solid')
        ws[letter + '1'].font = Font(color=colors.WHITE)\

    for (substitution, context), counts in data.items():
        ws.append([substitution,
                   context,
                   counts,
                   counts / sum(data.values())])

    wb.save(data_file)

    if args.title.lower() == 'none':
        title = None
    elif args.title == '':
        title = '{}\n($n={}$)'
        title = title.format(args.mutpos_file, sum(data.values()))
    else:
        title = args.title + '\n($n={}$)'.format(sum(data.values()))

    if args.ymax is not None:
        args.ymax = float(args.ymax)

    # Render plots and save
    if args.save == 'both' or args.save == 'ratio':
        spectrum_map(nrow=1, ncol=1,
                     heights=[[x * 100 for x in data.proportion]],
                     xlabels=[list(zip(*data.keys()))[1]],
                     labels=sorted(set(list(zip(*data.keys()))[0])),
                     titles=[title],
                     ylabel='% of Total Mutations\nNot Normalized',
                     y_max=args.ymax)
        plt.savefig(image_file1)

    if args.save == 'both' or args.save == 'total':
        spectrum_map(nrow=1, ncol=1,
                     heights=[list(data.values())],
                     xlabels=[list(zip(*data.keys()))[1]],
                     labels=sorted(set(list(zip(*data.keys()))[0])),
                     titles=[title],
                     ylabel='Number of Mutations',
                     y_max=args.ymax)
        plt.savefig(image_file2)


if __name__ == '__main__':
    main()